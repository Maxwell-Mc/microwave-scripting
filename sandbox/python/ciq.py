


def get_ciq_cascade(args):
    from openpyxl import load_workbook
    import openpyxl
    data = {}
    cascade = []
    status = False
    try:
        if args.filename:
            wb = load_workbook(args.filename)
            # Loop through sheets
            for ws in wb.worksheets:
                data[ws.title] = {}
                consecutive_col_none = 0
                consecutive_row_none = 0
                for col in range(1, ws.max_column + 1):
                    if ws[openpyxl.utils.get_column_letter(col) + '1' ].value == None:
                        consecutive_col_none = consecutive_col_none + 1
                    else:
                        consecutive_col_none = 0
                for row in range(1, ws.max_row + 1):
                    if ws[openpyxl.utils.get_column_letter(1) + str(row) ].value == None:
                        consecutive_row_none = consecutive_row_none + 1
                    else:
                        consecutive_row_none = 0
                for row in range(1, ws.max_row - consecutive_row_none):
                    data[ws.title][row]= {}
                    for col in range(1, ws.max_column - consecutive_col_none):
                        data[ws.title][row][ws[openpyxl.utils.get_column_letter(col) + '1'].value] = ws[openpyxl.utils.get_column_letter(col) + str(row) ].value

        for key, value in data['CIQs'].iteritems():
            if 'Site ID' in value.keys():
                if args.cascade in str(value['Site ID']):
                    cascade.append(value)
        status = True
    except:
        status = False
    
    return status, cascade

def main():
    import argparse
    from pprint import pprint

    parser = argparse.ArgumentParser(description='Process to load excel ciq documents')
    parser.add_argument('-f','--file', dest='filename', help='load to FILE', metavar='FILE')
    parser.add_argument('-c','--cascade', dest='cascade', help='search on the cascade', metavar='cascade')

    args = parser.parse_args()

    status, cascade = get_ciq_cascade(args)
    if status:
        pprint(cascade)
    else:
        pprint('Error: 02')
    


if __name__ == '__main__':
    main()