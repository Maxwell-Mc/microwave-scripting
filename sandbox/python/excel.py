from openpyxl import load_workbook
import openpyxl
import argparse
from pprint import pprint
import time


start = time.time()


parser = argparse.ArgumentParser(description='Process to load excel documents')

parser.add_argument('-f','--file', dest='filename', help='load to FILE', metavar='FILE')


args = parser.parse_args()

data = {}

if args.filename:
    wb = load_workbook(args.filename)
    # Loop through sheets
    for ws in wb.worksheets:
        data[ws.title] = {}
        consecutive_col_none = 0
        consecutive_row_none = 0
        for col in range(1, ws.max_column + 1):
            if ws[openpyxl.utils.get_column_letter(col) + '1' ].value == None:
                consecutive_col_none = consecutive_col_none + 1
            else:
                consecutive_col_none = 0
        for row in range(1, ws.max_row + 1):
            if ws[openpyxl.utils.get_column_letter(1) + str(row) ].value == None:
                consecutive_row_none = consecutive_row_none + 1
            else:
                consecutive_row_none = 0
        for row in range(1, ws.max_row - consecutive_row_none):
            data[ws.title][row]= {}
            for col in range(1, ws.max_column - consecutive_col_none):
                data[ws.title][row][ws[openpyxl.utils.get_column_letter(col) + '1'].value] = ws[openpyxl.utils.get_column_letter(col) + str(row) ].value
           
        # print ws
        # print 'max row = ' + str(ws.max_row)
        # print 'consecutive none row = ' + str(consecutive_row_none)
        # print 'max col = ' + str(ws.max_column)
        # print 'consecutive none col = ' + str(consecutive_col_none)
        # for row in range(1, ws.max_row + 1):
        #     data[ws.title][row]= {}
           
        #     # data[ws.title][ws[openpyxl.utils.get_column_letter('B') + str(row)].value] = {}
        #     for col in range(openpyxl.utils.column_index_from_string('A'), ws.max_column + 1):
        #         data[ws.title][row][ws[openpyxl.utils.get_column_letter(col) + '1'].value] = ws[openpyxl.utils.get_column_letter(col) + str(row) ].value
        
# pprint(data)  
end = time.time()
print (end - start)/60  + ' minutes'
