def get_ciq_cascade(args):
    from openpyxl import load_workbook
    import openpyxl
    data = {}
    cascade = []
    link = {}
    status = False
    try:
        if args.ciq:
            wb = load_workbook(args.ciq)
            # Loop through sheets
            for ws in wb.worksheets:
                data[ws.title] = {}
                consecutive_col_none = 0
                consecutive_row_none = 0
                for col in range(1, ws.max_column + 1):
                    if ws[openpyxl.utils.get_column_letter(col) + '1' ].value == None:
                        consecutive_col_none = consecutive_col_none + 1
                    else:
                        consecutive_col_none = 0
                for row in range(1, ws.max_row + 1):
                    if ws[openpyxl.utils.get_column_letter(1) + str(row) ].value == None:
                        consecutive_row_none = consecutive_row_none + 1
                    else:
                        consecutive_row_none = 0
                for row in range(1, ws.max_row - consecutive_row_none):
                    data[ws.title][row]= {}
                    for col in range(1, ws.max_column - consecutive_col_none): 
                        data[ws.title][row][ws[openpyxl.utils.get_column_letter(col) + '1'].value] = ws[openpyxl.utils.get_column_letter(col) + str(row) ].value
                    
        links = []
        for key, value in data['CIQs'].iteritems():
            if 'Site ID' in value.keys():
                if args.cascade in str(value['Site ID']):
                    cascade.append(value)
            if 'Link ID' in value.keys():
                if args.link_id in str(value['Link ID']):
                    links.append(value)
   
        link['a-side'] = {}
        for l in links:
            if args.cascade not in l['CASCADE']:
                link['a-side']['cascade'] = l['CASCADE']
        link[args.link_id] = links
        
        status = True
    except:
        status = False
    
    return status, cascade, link

def get_sheet_count(ws, h, data):
    import openpyxl
    consecutive_col_none = 0
    consecutive_row_none = 0
    # data = {}
    for col in range(1, ws.max_column + 1):
        if ws[openpyxl.utils.get_column_letter(col) + str(h) ].value == None:
            consecutive_col_none = consecutive_col_none + 1
        else:
            consecutive_col_none = 0
    for row in range(h, ws.max_row + 1):
        if ws[openpyxl.utils.get_column_letter(1) + str(row) ].value == None:
            consecutive_row_none = consecutive_row_none + 1
        else:
            consecutive_row_none = 0
    for row in range(1, ws.max_row - consecutive_row_none):
        data[ws.title][row]= {}
        for col in range(1, ws.max_column - consecutive_col_none):
            field = ws[openpyxl.utils.get_column_letter(col) + str(h)].value
            fvalue = ws[openpyxl.utils.get_column_letter(col) + str(row) ].value
            if field.strip() in data[ws.title][row]:
                data[ws.title][row][field.strip() + '_1'] = fvalue
            else:
                data[ws.title][row][field.strip()] = fvalue
    return data[ws.title]

def get_ip_plan_cascade(args, current_data):
    from openpyxl import load_workbook
    import openpyxl
    data = {}
    cascade = {}
    status = False
    try:
        if args.ip_plan:
            wb = load_workbook(args.ip_plan)
            # Loop through sheets
            for ws in wb.worksheets:
                data[ws.title] = {}
                if '3G IP Plan' in ws.title:
                    data[ws.title] = get_sheet_count(ws, 3, data)
                elif '4G IP Plan' in ws.title:
                    data[ws.title] = get_sheet_count(ws, 2, data)
                else:
                    data[ws.title] = get_sheet_count(ws, 1, data)

        ip_plan_3g = []
        for key, value in data['3G IP Plan'].iteritems():
            field = "Cascade_ID\ncell site-id"
            if field in value.keys():
                if args.cascade in str(value[field]):
                    ip_plan_3g.append(value)
        cascade['3G IP Plan'] = ip_plan_3g
        ip_plan_4g = []
        for key, value in data['4G IP Plan'].iteritems():
            field = "Cascade_ID\ncell site-id"
            if field in value.keys():
                if args.cascade in str(value[field]):
                    ip_plan_4g.append(value)

        cascade['4G IP Plan'] = ip_plan_4g

        ip_plan_4g_aside = []
        ciq_aside = []
        cascade['a-side'] = {}
        cascade['a-side']['cascade'] = current_data['links']['a-side']['cascade'] 
        for item in current_data['links'][args.link_id]:
            if cascade['a-side']['cascade'] in str(item['CASCADE']):
                ciq_aside.append(item)

        cascade['a-side']['ciq'] = ciq_aside

        for key, value in data['4G IP Plan'].iteritems():
            field = "Cascade_ID\ncell site-id"
            if field in value.keys():
                if cascade['a-side']['cascade'] in str(value[field]):
                    ip_plan_4g_aside.append(value)

        cascade['a-side']['4G IP Plan'] = ip_plan_4g_aside
    except:
        pass
    return True, cascade

def render(tpl_path, context):
    import os
    import jinja2

    path, filename = os.path.split(tpl_path)
    return jinja2.Environment(
        loader=jinja2.FileSystemLoader(path or './')
    ).get_template(filename).render(context)

def main():
    import argparse
    from pprint import pprint
    import time

    start = time.time()

    data ={}

    parser = argparse.ArgumentParser(description='Process to load excel ciq documents')
    parser.add_argument('-q','--ciq', dest='ciq', help='This option is the destination file for CIQ document', metavar='FILE')
    parser.add_argument('-p','--ip_plan', dest='ip_plan', help='This option is the destination file for sites IP Plan document', metavar='FILE')
    parser.add_argument('-c','--cascade', dest='cascade', help='This option is to search on the cascade in question', metavar='string')
    parser.add_argument('-qsi','--qscope_interface', dest='qscope_interface', help='This option is to set the Qscope Interface location on the IPA', metavar='string')
    parser.add_argument('-l','--link_id', dest='link_id', help='This option is to set the link id for the site you are building', metavar='string')
    

    args = parser.parse_args()

    data['qscope'] = args.qscope_interface
    data['link_id'] = args.link_id
    # Data Collection at cascaded level
    ciq_status, cascade_ciq, cascade_link = get_ciq_cascade(args)
    if ciq_status:
        data['ciq'] = cascade_ciq
        data['links'] = cascade_link
    else:
        data['ciq'] ='Error: 02'
    
    plan_status, cascade_plan = get_ip_plan_cascade(args, data)
    if plan_status:
        data['ip_plan'] = cascade_plan

    pprint(data)
    # fd = open("data.txt","w")
    # fd.write(str(data))
    # fd.close()

    result1 = render('resources/templates/straight_to_fiber_ipa1.j2', data)

    fs1 = open("ipa1.txt","w")
    fs1.write(result1)
    fs1.close()

    result2 = render('resources/templates/straight_to_fiber_ipa2.j2', data)

    fs2 = open("ipa2.txt","w")
    fs2.write(result2)
    fs2.close()
    # pprint(result)

    end = time.time()
    total = (end - start)/60
    minute = round(total)
    arr_nin = str(minute).split('.')
    sec = round(((total - minute) * 60)/100,2)
    arr_sec = str(sec).split('.')

    s = ":"
    print s.join([arr_nin[0], arr_sec[1]])
    
    


if __name__ == '__main__':
    main()